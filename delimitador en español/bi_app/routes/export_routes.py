from flask import Blueprint, request, jsonify, send_file
import pandas as pd
import os, sys
from datetime import datetime
sys.path.insert(0, os.path.dirname(os.path.dirname(__file__)))
from modules.data_manager import get_clean, get_original, has_data, get_numeric_cols, get_categorical_cols

export_bp = Blueprint('export', __name__)
EXPORT_FOLDER = os.path.join(os.path.dirname(os.path.dirname(__file__)), 'exports')

@export_bp.route('/excel', methods=['POST'])
def export_excel():
    if not has_data(): return jsonify({'error': 'No hay dataset'}), 400
    df = get_clean(); orig = get_original()
    num_cols = get_numeric_cols(df); cat_cols = get_categorical_cols(df)
    os.makedirs(EXPORT_FOLDER, exist_ok=True)
    ts = datetime.now().strftime('%Y%m%d_%H%M%S')
    path = os.path.join(EXPORT_FOLDER, f'BI_{ts}.xlsx')
    try:
        with pd.ExcelWriter(path, engine='openpyxl') as writer:
            df.to_excel(writer, sheet_name='Datos_Limpios', index=False)
            orig.to_excel(writer, sheet_name='Datos_Originales', index=False)
            if num_cols:
                st = df[num_cols].describe(percentiles=[.1,.25,.5,.75,.9]).T
                try:
                    st['asimetria'] = df[num_cols].skew()
                    st['curtosis'] = df[num_cols].kurtosis()
                except Exception: pass
                st.reset_index().rename(columns={'index':'variable'}).to_excel(writer, sheet_name='Estadísticas', index=False)
            if len(num_cols) >= 2:
                df[num_cols].corr().round(4).to_excel(writer, sheet_name='Correlaciones')
            if cat_cols:
                rows = []
                for col in cat_cols:
                    vc = df[col].astype(str).value_counts()
                    for cat, cnt in vc.items():
                        rows.append({'Columna':col,'Categoría':str(cat),'Frecuencia':int(cnt),'%':round(cnt/max(len(df),1)*100,2)})
                pd.DataFrame(rows).to_excel(writer, sheet_name='Frecuencias', index=False)
            miss = df.isnull().sum().reset_index(); miss.columns=['Columna','Faltantes']
            miss['%'] = (miss['Faltantes']/max(len(df),1)*100).round(2)
            miss.to_excel(writer, sheet_name='Faltantes', index=False)
            try:
                from openpyxl.styles import PatternFill, Font, Alignment
                hf = PatternFill(start_color='1E3A5F', end_color='1E3A5F', fill_type='solid')
                hfont = Font(color='FFFFFF', bold=True)
                for ws in writer.book.worksheets:
                    for cell in ws[1]:
                        cell.fill = hf; cell.font = hfont
                        cell.alignment = Alignment(horizontal='center')
                    for col_cells in ws.columns:
                        mx = max((len(str(c.value or '')) for c in col_cells), default=8)
                        ws.column_dimensions[col_cells[0].column_letter].width = min(mx+3, 40)
            except Exception: pass
        return send_file(path, as_attachment=True, download_name=f'BI_{ts}.xlsx',
                         mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@export_bp.route('/pdf', methods=['POST'])
def export_pdf():
    if not has_data(): return jsonify({'error': 'No hay dataset'}), 400
    df = get_clean(); num_cols = get_numeric_cols(df); cat_cols = get_categorical_cols(df)
    os.makedirs(EXPORT_FOLDER, exist_ok=True)
    ts = datetime.now().strftime('%Y%m%d_%H%M%S')
    path = os.path.join(EXPORT_FOLDER, f'BI_{ts}.pdf')
    try:
        from reportlab.lib import colors
        from reportlab.lib.pagesizes import A4
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.lib.units import cm
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, HRFlowable
        from reportlab.lib.enums import TA_CENTER

        doc = SimpleDocTemplate(path, pagesize=A4, topMargin=1.5*cm, bottomMargin=1.5*cm, leftMargin=2*cm, rightMargin=2*cm)
        styles = getSampleStyleSheet()
        S = lambda name, **kw: ParagraphStyle(name, parent=styles['Normal'], **kw)
        title_st = S('T', fontSize=20, textColor=colors.HexColor('#0F172A'), alignment=TA_CENTER, spaceAfter=6, fontName='Helvetica-Bold')
        h1_st = S('H1', fontSize=13, textColor=colors.HexColor('#1e40af'), spaceBefore=12, spaceAfter=4, fontName='Helvetica-Bold')
        body_st = S('B', fontSize=10, spaceAfter=4)

        def mktable(data, col_widths, header_color='#1e40af', row_colors=None):
            t = Table(data, colWidths=col_widths)
            rc = row_colors or [colors.HexColor('#f8fafc'), colors.white]
            t.setStyle(TableStyle([
                ('BACKGROUND',(0,0),(-1,0),colors.HexColor(header_color)),
                ('TEXTCOLOR',(0,0),(-1,0),colors.white),
                ('FONTNAME',(0,0),(-1,0),'Helvetica-Bold'),
                ('FONTSIZE',(0,0),(-1,-1),9),
                ('ROWBACKGROUNDS',(0,1),(-1,-1),rc),
                ('GRID',(0,0),(-1,-1),0.5,colors.HexColor('#e2e8f0')),
                ('ALIGN',(1,0),(-1,-1),'CENTER'),
                ('PADDING',(0,0),(-1,-1),5),
            ]))
            return t

        story = [Spacer(1,0.5*cm), Paragraph("Business Intelligence Report", title_st),
                 Paragraph(f"Generado: {datetime.now().strftime('%d/%m/%Y %H:%M')}", body_st),
                 HRFlowable(width='100%', thickness=2, color=colors.HexColor('#1e40af')), Spacer(1,0.4*cm)]

        miss_total = int(df.isnull().sum().sum()); tc = df.shape[0]*df.shape[1]
        story.append(Paragraph("1. Resumen del Dataset", h1_st))
        story.append(mktable([['Métrica','Valor'],['Filas',str(len(df))],['Columnas',str(len(df.columns))],
                               ['Numéricas',str(len(num_cols))],['Categóricas',str(len(cat_cols))],
                               ['Faltantes',str(miss_total)],['Calidad',f"{round((1-miss_total/max(tc,1))*100,1)}%"]],
                              [9*cm,5*cm]))
        story.append(Spacer(1,0.4*cm))

        if num_cols:
            story.append(Paragraph("2. Estadísticas Descriptivas", h1_st))
            hdr = ['Variable','Media','Mediana','Std','Mín','Máx','Asimetría']
            rows = [hdr] + [[col[:18],f"{df[col].mean():.2f}",f"{df[col].median():.2f}",
                              f"{df[col].std():.2f}",f"{df[col].min():.2f}",f"{df[col].max():.2f}",
                              f"{df[col].skew():.3f}"] for col in num_cols[:12]]
            story.append(mktable(rows, [4.5*cm,2.5*cm,2.5*cm,2.5*cm,2.5*cm,2.5*cm,2.5*cm], '#7c3aed',
                                  [colors.HexColor('#f5f3ff'), colors.white]))
            story.append(Spacer(1,0.4*cm))

        if cat_cols:
            story.append(Paragraph("3. Variables Categóricas", h1_st))
            for col in cat_cols[:4]:
                story.append(Paragraph(f"• {col}", body_st))
                vc = df[col].astype(str).value_counts().head(5)
                rows = [['Categoría','Frec.','%']] + [[str(k)[:28],str(int(v)),f"{round(v/max(len(df),1)*100,1)}%"] for k,v in vc.items()]
                story.append(mktable(rows, [8*cm,4*cm,4*cm], '#059669',
                                      [colors.HexColor('#f0fdf4'), colors.white]))
                story.append(Spacer(1,0.2*cm))

        story += [Spacer(1,0.8*cm), HRFlowable(width='100%',thickness=1,color=colors.HexColor('#e2e8f0')),
                  Paragraph("Generado por BI Platform", S('f',fontSize=8,textColor=colors.grey,alignment=TA_CENTER))]
        doc.build(story)
        return send_file(path, as_attachment=True, download_name=f'BI_{ts}.pdf', mimetype='application/pdf')
    except Exception as e:
        import traceback
        return jsonify({'error': str(e), 'detail': traceback.format_exc()}), 500
